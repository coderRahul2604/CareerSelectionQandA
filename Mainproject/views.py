from django.shortcuts import render, HttpResponse, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.db.models import Q
from Question.models import Question, Answer, Comment,Like, Dislike
from Career.models import Career, Career2
from .models import Profile
from django.http import HttpResponse
from django.core.mail import send_mail
from django.conf import settings
import uuid

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Create your views here.
def index(request):
    if 'career' in request.GET:
        career = request.GET['career']
        career_query = Q(tags__icontains=career)
        career_results = Career.objects.filter(career_query)
        career2_results = Career2.objects.filter(career_query)
        results = list(career_results) + list(career2_results)
        return render(request, 'search.html', {'career': results})
    
    return render(request, 'index.html')

def About(request):
    return render(request, 'About/about.html')

def Signup(request):
    if request.method=="POST":
        username=request.POST['username']
        email=request.POST['email']
        pass1=request.POST['pass1']
        pass2=request.POST['pass2']

        if pass1 != pass2:
            messages.error(request, "Password should be Same")
            return redirect('index')
        
        if len(username) < 3 and len(username)> 15:
            messages.error(request, "Username should be contain 4-15 charaters only")
            return redirect('index')
        
        if not username.isalnum():
            messages.error(request, "Username should be contain letters and numbers")
            return redirect('index')

        myuser=User.objects.create_user(username, email, pass1,)
        myuser.save()
        messages.success(request, "Your account has been successfully created")
        return render(request, 'index.html')

    else:
        return HttpResponse("404 error")

def Login(request):
    if request.method=="POST":
        loginusername = request.POST['loginusername']
        loginpassword = request.POST['loginpassword']
        user = authenticate(request, username=loginusername, password=loginpassword)
    if user is not None:
        login(request, user)
        messages.success(request, "Successfully loged in ")
        return render(request, 'index.html')
   
    else:
        messages.error(request, 'Invalid credentional, please try again')
        return render(request, 'index.html')
       
def Logout(request):
    logout(request)
    messages.success(request, "Successfully loged out")
    return render(request, 'index.html')

def forgotpassword(request):
    return render(request, 'Login/forgotpass.html')

def ikigai(request):
    if request.user.is_authenticated:
        return render(request, 'ikigai.html')
    else:
        return render(request, 'Login/Signup.html')
    
def profile(request, username):
    user = User.objects.get(username=username)
    questions = Question.objects.filter(user=user)
    answers = Answer.objects.filter(user=user)
    comments = Comment.objects.filter(user=user)
    likes = Like.objects.filter(user=user)
    dislikes = Dislike.objects.filter(user=user)
    context= {'user': user, 'questions': questions, 'answers': answers,'comments':comments,'likes':likes, 'dislikes':dislikes}
    return render(request, 'Profile/profile.html', context)

def send_forgot_passwod_mail(reuest, token):
    pass

def forgotpassword(request):
    return render(request , 'forget-password.html')

def resetpassword(request, token):
    return render(request, 'Login/resetpass.html')

def report(request):
    if request.user.is_authenticated and request.user.is_superuser:
        return render(request, 'report.html')
    else:
        return HttpResponse("Error")
    
def Question_report(request):
    if request.user.is_authenticated and request.user.is_superuser:
        # Create a response object with the content type set as Excel
        response = HttpResponse(content_type='application/ms-excel')

        # Set the filename for the Excel file
        response['Content-Disposition'] = 'attachment; filename="question_report.xlsx"'

        # Retrieve data from your database table
        queryset = Question.objects.all()

        # Create a new workbook object
        workbook = Workbook()

        # Get the active worksheet
        worksheet = workbook.active

        # Add column headings and set font style
        worksheet['A1'].value = 'Username'
        worksheet['A1'].font = Font(bold=True) 
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['B1'].value = 'Question'
        worksheet['B1'].font = Font(bold=True)
        worksheet['B1'].alignment = Alignment(horizontal='center')
        worksheet['C1'].value = 'Date & Time'
        worksheet['C1'].font = Font(bold=True)
        worksheet['C1'].alignment = Alignment(horizontal='center')

        # Add data to the worksheet
        row_num = 2

        for user in queryset:
            # Convert the datetime object to a string with timezone information removed
            date_string = user.add_time.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S')
            worksheet.cell(row=row_num, column=1, value=str(user.user))
            worksheet.cell(row=row_num, column=2, value=str(user.title))
            worksheet.cell(row=row_num, column=3, value=date_string)

            # Wrap column data
            for column in range(1, 4):
                worksheet.cell(row=row_num, column=column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


            row_num += 1

        # Save the workbook
        workbook.save(response)

        # Return the response object
        return response
    else:
        return HttpResponse("Error")
    
def Answer_report(request):
    if request.user.is_authenticated and request.user.is_superuser:
        # Create a response object with the content type set as Excel
         response = HttpResponse(content_type='application/ms-excel')

         # Set the filename for the Excel file
         response['Content-Disposition'] = 'attachment; filename="answer_report.xlsx"'

         # Retrieve data from your database table
         queryset = Answer.objects.all()

         # Create a new workbook object
         workbook = Workbook()

         # Get the active worksheet
         worksheet = workbook.active

         # Add column headings
         bold_font = Font(bold=True)
         worksheet['A1'] = 'Username'
         worksheet['A1'].font = bold_font
         worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

         worksheet['B1'] = 'Question'
         worksheet['B1'].font = bold_font
         worksheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

         worksheet['C1'] = 'Answer'
         worksheet['C1'].font = bold_font
         worksheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

         worksheet['D1'] = 'Date & Time'
         worksheet['D1'].font = bold_font
         worksheet['D1'].alignment = Alignment(horizontal='center', vertical='center')

         

         # Add data to the worksheet
         row_num = 2

         for user in queryset:
             # Convert the datetime object to a string with timezone information removed
            date_string = user.add_time.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S')
            worksheet.cell(row=row_num, column=1, value=str(user.user))
            worksheet.cell(row=row_num, column=2, value=str(user.question))
            worksheet.cell(row=row_num, column=3, value=str(user.detail))
            worksheet.cell(row=row_num, column=4, value=date_string)
            for column in range(1, 5):
                worksheet.cell(row=row_num, column=column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
             

            row_num += 1

         # Save the workbook
         workbook.save(response)

         # Return the response object
         return response
    else:
        return HttpResponse("Error")

def Comment_report(request):
    if request.user.is_authenticated and request.user.is_superuser:
        # Create a response object with the content type set as Excel
        response = HttpResponse(content_type='application/ms-excel')

        # Set the filename for the Excel file
        response['Content-Disposition'] = 'attachment; filename="comment_report.xlsx"'

        # Retrieve data from your database table
        queryset = Comment.objects.all()

        # Create a new workbook object
        workbook = Workbook()

        # Get the active worksheet
        worksheet = workbook.active

        # Add column headings
        bold_font = Font(bold=True)
        worksheet['A1'] = 'Username'
        worksheet['A1'].font = bold_font
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['B1'] = 'Answer'
        worksheet['B1'].font = bold_font
        worksheet['B1'].alignment = Alignment(horizontal='center')
        worksheet['C1'] = 'Comment'
        worksheet['C1'].font = bold_font
        worksheet['C1'].alignment = Alignment(horizontal='center')
        worksheet['D1'] = 'Date & Time'
        worksheet['D1'].font = bold_font
        worksheet['D1'].alignment = Alignment(horizontal='center')


        # Add data to the worksheet
        row_num = 2

        for user in queryset:
            # Convert the datetime object to a string with timezone information removed
            date_string = user.add_time.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S')
            worksheet.cell(row=row_num, column=1, value=str(user.user))
            worksheet.cell(row=row_num, column=2, value=str(user.answer))
            worksheet.cell(row=row_num, column=3, value=str(user.comment))
            worksheet.cell(row=row_num, column=4, value=date_string)
            for column in range(1, 5):
                worksheet.cell(row=row_num, column=column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            row_num += 1

        # Save the workbook
        workbook.save(response)

        # Return the response object
        return response
    else:
        return HttpResponse("Error")

def Profile_report(request):
    if request.user.is_authenticated and request.user.is_superuser:
        # Create a response object with the content type set as Excel
        response = HttpResponse(content_type='application/ms-excel')

        # Set the filename for the Excel file
        response['Content-Disposition'] = 'attachment; filename="profile_report.xlsx"'

        # Retrieve data from your database table
        queryset = User.objects.all()

        # Create a new workbook object
        workbook = Workbook()

        # Get the active worksheet
        worksheet = workbook.active

        # Add column headings
        bold_font = Font(bold=True)
        worksheet['A1'] = 'Username'
        worksheet['A1'].font = bold_font
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['B1'] = 'Email'
        worksheet['B1'].font = bold_font
        worksheet['B1'].alignment = Alignment(horizontal='center')
        worksheet['C1'] = 'Date & Time'
        worksheet['C1'].font = bold_font
        worksheet['C1'].alignment = Alignment(horizontal='center')


        # Add data to the worksheet
        row_num = 2

        for user in queryset:
            # Convert the datetime object to a string with timezone information removed
            date_string = user.date_joined.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S')
            worksheet.cell(row=row_num, column=1, value=str(user.username))
            worksheet.cell(row=row_num, column=2, value=str(user.email))
            worksheet.cell(row=row_num, column=3, value=date_string)
            for column in range(1, 4):
                worksheet.cell(row=row_num, column=column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            row_num += 1

        # Save the workbook
        workbook.save(response)

        # Return the response object
        return response
    else:
        return HttpResponse("Error")

def Like_report(request):
    if request.user.is_authenticated and request.user.is_superuser:
        # Create a response object with the content type set as Excel
        response = HttpResponse(content_type='application/ms-excel')

        # Set the filename for the Excel file
        response['Content-Disposition'] = 'attachment; filename="Like_report.xlsx"'

        # Retrieve data from your database table
        queryset = Like.objects.all()

        # Create a new workbook object
        workbook = Workbook()

        # Get the active worksheet
        worksheet = workbook.active

        # Add column headings
        bold_font = Font(bold=True)
        worksheet['A1'] = 'Answer'
        worksheet['A1'].font = bold_font
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['B1'] = 'Username'
        worksheet['B1'].font = bold_font
        worksheet['B1'].alignment = Alignment(horizontal='center')
        worksheet['C1'] = 'Date & Time'
        worksheet['C1'].font = bold_font
        worksheet['C1'].alignment = Alignment(horizontal='center')


        # Add data to the worksheet
        row_num = 2

        for user in queryset:
            # Convert the datetime object to a string with timezone information removed
            date_string = user.add_time.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S')
            worksheet.cell(row=row_num, column=1, value=str(user.answer))
            worksheet.cell(row=row_num, column=2, value=str(user.user))
            worksheet.cell(row=row_num, column=3, value=date_string)
            for column in range(1, 4):
                worksheet.cell(row=row_num, column=column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            row_num += 1

        # Save the workbook
        workbook.save(response)

        # Return the response object
        return response
    else:
        return HttpResponse("Error")

def Dislike_report(request):
    if request.user.is_authenticated and request.user.is_superuser:
        # Create a response object with the content type set as Excel
        response = HttpResponse(content_type='application/ms-excel')

        # Set the filename for the Excel file
        response['Content-Disposition'] = 'attachment; filename="Dislike_report.xlsx"'

        # Retrieve data from your database table
        queryset = Dislike.objects.all()

        # Create a new workbook object
        workbook = Workbook()

        # Get the active worksheet
        worksheet = workbook.active

        # Add column headings
        bold_font = Font(bold=True)
        worksheet['A1'] = 'Answer'
        worksheet['A1'].font = bold_font
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['B1'] = 'Username'
        worksheet['B1'].font = bold_font
        worksheet['B1'].alignment = Alignment(horizontal='center')
        worksheet['C1'] = 'Date & Time'
        worksheet['C1'].font = bold_font
        worksheet['C1'].alignment = Alignment(horizontal='center')


        # Add data to the worksheet
        row_num = 2

        for user in queryset:
            # Convert the datetime object to a string with timezone information removed
            date_string = user.add_time.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S')
            worksheet.cell(row=row_num, column=1, value=str(user.answer))
            worksheet.cell(row=row_num, column=2, value=str(user.user))
            worksheet.cell(row=row_num, column=3, value=date_string)
            for column in range(1, 4):
                worksheet.cell(row=row_num, column=column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            row_num += 1

        # Save the workbook
        workbook.save(response)

        # Return the response object
        return response
    else:
        return HttpResponse("Error")

def Askquestion(request):
    if request.user.is_authenticated:
        if request.method=='POST':
            que = Question()
            que.title = request.POST.get('quetitle')
            que.details = request.POST.get('details')
            que.tags = request.POST.get('tags')
            que.user = request.user
            que.save()
            comment_date_time = que.add_time
            return render(request, 'Askquestion.html')
    else:
        return render(request, 'Login/Signup.html')
    return render(request, 'Askquestion.html')

